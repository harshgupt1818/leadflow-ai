# exporter.py
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def export_to_excel(leads: list, filename: str = "leads_output.xlsx"):
    
    df = pd.DataFrame(leads)
    
    column_order = [
        "name", "phone", "website", "address",
        "rating", "score", "status",
        "reason", "outreach_message"
    ]
    
    existing_cols = [c for c in column_order if c in df.columns]
    df = df[existing_cols]
    
    if 'score' in df.columns:
        df = df.sort_values('score', ascending=False)
    
    # Duplicates hatao
    if 'name' in df.columns:
        df = df.drop_duplicates(subset=['name'])
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Leads')
        
        workbook = writer.book
        worksheet = writer.sheets['Leads']
        
        # Header styling
        header_fill = PatternFill(
            start_color="1F4E79",
            end_color="1F4E79",
            fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Row coloring
        for row in range(2, len(df) + 2):
            status_col = existing_cols.index('status') + 1 if 'status' in existing_cols else None
            
            if status_col:
                status = worksheet.cell(row=row, column=status_col).value
                
                if status == "Hot":
                    color = "FFD7D7"
                elif status == "Warm":
                    color = "FFF3CD"
                else:
                    color = "E8F5E9"
                
                fill = PatternFill(
                    start_color=color,
                    end_color=color,
                    fill_type="solid"
                )
                
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row, column=col).fill = fill
        
        # Column width
        for col in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = 25
    
    print(f"✅ Excel saved: {filename}")
    return filename